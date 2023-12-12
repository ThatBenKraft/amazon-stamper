from typing import Any

import openpyxl

FILENAME = "stamper_data.xlsx"
SHEETNAME = "stamper_data"


class Cells:
    RUNNING = "A1"
    CODE = "A2"
    STOP = "A3"


def print_sheet(sheet) -> None:
    print([cell[0].value for cell in sheet])


def set_cell(cell: str, value) -> None:
    """
    Sets specified cell in worksheet to value.
    """
    workbook, sheet = _get_workbook()
    sheet[cell] = value
    workbook.save(FILENAME)


def get_cell(cell: str) -> Any:
    """
    Gets value from specified cell in worksheet.
    """
    _, sheet = _get_workbook()
    return sheet[cell].value


def _get_workbook() -> tuple[openpyxl.Workbook, openpyxl.worksheet]:  # type: ignore
    """
    Returns workbook and worksheet, creates new if they do not exist.
    """
    try:
        workbook = openpyxl.load_workbook(FILENAME)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.create_sheet(SHEETNAME)
        workbook.save(FILENAME)
    return workbook, workbook[SHEETNAME]


if __name__ == "__main__":
    # Test actions
    set_cell(Cells.RUNNING, False)

    set_cell(Cells.CODE, "B00B")

    print(get_cell(Cells.CODE))

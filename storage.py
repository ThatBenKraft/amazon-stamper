#!/usr/bin/env python
"""
## Storage
Allows for storage of stamper data, such as code and running status.
"""

from typing import Any

from openpyxl import Workbook, load_workbook

__author__ = "Ben Kraft"
__copyright__ = "None"
__credits__ = "Ben Kraft"
__license__ = "Apache"
__version__ = "0.0.1"
__maintainer__ = "Ben Kraft"
__email__ = "ben.kraft@rcn.com"
__status__ = "Prototype"

FILENAME = "stamper_data.xlsx"
SHEETNAME = "stamper_data"


class Cells:
    """
    Establishes cells for storage.
    """

    RUNNING = "A1"
    CODE = "A2"
    STOP = "A3"


def print_sheet(sheet: Any) -> None:
    """
    Prints values in specified sheet.
    """
    print([cell[0].value for cell in sheet])


def set_cell(cell: str, value) -> None:
    """
    Sets specified cell in worksheet to value.
    """
    # Gets workbook and sheet from file
    workbook, sheet = _get_workbook()
    # Sets cell to value
    sheet[cell] = value
    # Saves workbook to file
    workbook.save(FILENAME)


def get_cell(cell: str) -> Any:
    """
    Gets value from specified cell in worksheet.
    """
    # Gets sheet from file
    _, sheet = _get_workbook()
    # Returns cell value
    return sheet[cell].value


def _get_workbook() -> tuple[Workbook, Any]:  # type: ignore
    """
    Returns workbook and worksheet, creates new if they do not exist.
    """
    try:
        # Loads workbook from file
        workbook = load_workbook(FILENAME)
    except FileNotFoundError:
        # Creates workbook and worksheet
        workbook = Workbook()
        workbook.create_sheet(SHEETNAME)
        # Saves new workbook
        workbook.save(FILENAME)
    # Returns workbook and worksheet objects
    return workbook, workbook[SHEETNAME]


if __name__ == "__main__":
    # Prints out running status and code
    print(f"Running: {get_cell(Cells.RUNNING)}")
    print(f"Code: {get_cell(Cells.CODE)}")
